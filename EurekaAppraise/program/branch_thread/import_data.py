#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sqlite3
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5 import QtCore
from ..initialize_setting import EDIT_TABLE, SPECIAL_TABLE
from ..asset_method.edit_table.edit_model import EditModel
from ..asset_method.input_table.input_model import InputModel


class ImportDataThread(QtCore.QThread):
    conn: sqlite3.Connection
    wb: Workbook
    logPrinter = QtCore.pyqtSignal(str)
    errorPrinter = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal(bool)

    def __init__(self, file: str, path: str, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.file = file
        self.path = path

    def run(self):
        self.logPrinter.emit(self.tr('Load File'))
        self.wb = load_workbook(self.path, data_only=True)
        self.conn = sqlite3.connect(self.file)
        for ws_name in self.wb.sheetnames:
            if ws_name in EDIT_TABLE:
                self.paste_edit_table(ws_name)
            elif ws_name in SPECIAL_TABLE:
                self.paste_special_table(ws_name)
        self.wb.close()
        self.conn.close()
        self.logPrinter.emit('Import data finished')
        self.finished.emit(True)

    def paste_edit_table(self, worksheet_name):
        self.logPrinter.emit(self.tr('Import ') + worksheet_name)
        model = EditModel(self.conn, worksheet_name)
        ws: Worksheet = self.wb[worksheet_name]
        temp = tuple(ws.values)
        end = temp[6].index('备注')
        data = [[str(cell) if cell else '' for cell in row[1: end]] for row in temp[7:] if any(row)]
        data_text = '\n'.join(['\t'.join(row) for row in data])
        model.paste_range([model.createIndex(0, 0)], data_text)

    def paste_special_table(self, worksheet_name):
        self.logPrinter.emit(self.tr('Import ') + worksheet_name)
        model = InputModel(self.conn, worksheet_name)
        ws: Worksheet = self.wb[worksheet_name]
        if worksheet_name == '表7':
            temp = tuple(ws['C8:D16'])
        else:
            temp = tuple(ws['C8:D38'])
        data = [[str(cell.value) if cell else '' for cell in row] for row in temp]
        data_text = '\n'.join(['\t'.join(row) for row in data])
        model.paste_range([model.createIndex(0, 1)], data_text)
